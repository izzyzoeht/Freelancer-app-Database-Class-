"""
/api/revenue - rev summary for model
"""
from flask import Blueprint, jsonify, send_file
from routes._helpers import get_db, login_required, current_user_type
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.axis import ChartLines
from io import BytesIO
from datetime import datetime

revenue = Blueprint('revenue', __name__)

@revenue.route('/summary', methods=['GET'])
@login_required
def summary():
    if current_user_type() != 'Admin':
        return jsonify({'error': 'Only admins can view the revenue summary'}), 403

    db = get_db()
    cursor = db.cursor(dictionary=True)

    cursor.execute(
        "SELECT COALESCE(SUM(amount), 0) AS total_payment_volume FROM payments WHERE status='paid'"
    )
    total_payment_volume = cursor.fetchone()['total_payment_volume']

    cursor.execute("SELECT COALESCE(SUM(fee_amount), 0) AS total_platform_fees FROM platform_fees")
    total_platform_fees = cursor.fetchone()['total_platform_fees']

    cursor.execute(
        """SELECT COUNT(*) AS active_subscriptions,
                  COALESCE(SUM(price_at_purchase), 0) AS monthly_subscription_revenue
             FROM subscriptions
            WHERE status='active'"""
    )
    subs = cursor.fetchone()

    cursor.execute(
        """SELECT rs.stream_name,
                  COALESCE(SUM(pf.fee_amount), 0) AS revenue
             FROM revenue_streams rs
             LEFT JOIN platform_fees pf ON pf.revenue_stream_id = rs.revenue_stream_id
            GROUP BY rs.revenue_stream_id, rs.stream_name
            ORDER BY rs.stream_name"""
    )
    streams = cursor.fetchall()

    cursor.close(); db.close()

    total_estimated_revenue = total_platform_fees + subs['monthly_subscription_revenue']
    return jsonify({
        'total_payment_volume': float(total_payment_volume),
        'total_platform_fees': float(total_platform_fees),
        'active_subscriptions': int(subs['active_subscriptions']),
        'monthly_subscription_revenue': float(subs['monthly_subscription_revenue']),
        'total_estimated_revenue': float(total_estimated_revenue),
        'streams': [
            {'stream_name': r['stream_name'], 'revenue': float(r['revenue'])}
            for r in streams
        ],
    }), 200

@revenue.route('/export', methods=['GET'])
@login_required
def export_excel():
    if current_user_type() != 'Admin':
        return jsonify({'error': 'Only admins can export the revenue report'}), 403

    db = get_db()
    cursor = db.cursor(dictionary=True)

    # ---- Summary metrics ----
    cursor.execute("SELECT COALESCE(SUM(amount), 0) AS total FROM payments WHERE status='paid'")
    total_payment_volume = float(cursor.fetchone()['total'])

    cursor.execute("SELECT COALESCE(SUM(fee_amount), 0) AS total FROM platform_fees")
    total_platform_fees = float(cursor.fetchone()['total'])

    cursor.execute("""
        SELECT COUNT(*) AS active_subs,
               COALESCE(SUM(price_at_purchase), 0) AS mrr
        FROM subscriptions WHERE status='active'
    """)
    sub_row = cursor.fetchone()
    active_subscriptions = int(sub_row['active_subs'])
    monthly_subscription_revenue = float(sub_row['mrr'])

    total_estimated_revenue = total_platform_fees + monthly_subscription_revenue

    # ---- Monthly revenue ----
    cursor.execute("""
        SELECT DATE_FORMAT(p.paid_at, '%Y-%m') AS month,
               COUNT(*) AS num_payments,
               COALESCE(SUM(p.amount), 0) AS total_revenue,
               COALESCE(SUM(pf.fee_amount), 0) AS platform_fees,
               COUNT(DISTINCT b.user_id) AS unique_employers,
               COUNT(DISTINCT b.tradesperson_id) AS unique_tradespeople
        FROM payments p
        JOIN bookings b ON p.booking_id = b.booking_id
        LEFT JOIN platform_fees pf ON pf.payment_id = p.payment_id
        WHERE p.status='paid'
        GROUP BY month
        ORDER BY month
    """)
    monthly_data = cursor.fetchall()

    # ---- Revenue by city ----
    cursor.execute("""
        SELECT b.city,
               COUNT(*) AS bookings,
               COALESCE(SUM(p.amount), 0) AS revenue,
               COALESCE(SUM(pf.fee_amount), 0) AS platform_fees,
               COUNT(DISTINCT b.tradesperson_id) AS active_tradespeople
        FROM bookings b
        JOIN payments p ON b.booking_id = p.booking_id
        LEFT JOIN platform_fees pf ON pf.payment_id = p.payment_id
        WHERE p.status='paid'
        GROUP BY b.city
        ORDER BY revenue DESC
    """)
    city_data = cursor.fetchall()

    # ---- Top tradespeople ----
    cursor.execute("""
        SELECT CONCAT(u.first_name, ' ', u.last_name) AS name,
               t.trade_category,
               u.city,
               COUNT(DISTINCT b.booking_id) AS total_bookings,
               COALESCE(SUM(p.amount), 0) AS total_revenue,
               COALESCE(t.avg_rating, 0) AS avg_rating
        FROM tradespeople t
        JOIN users u ON t.user_id = u.user_id
        LEFT JOIN bookings b ON b.tradesperson_id = t.tradesperson_id
        LEFT JOIN payments p ON p.booking_id = b.booking_id AND p.status='paid'
        GROUP BY t.tradesperson_id, name, t.trade_category, u.city, t.avg_rating
        HAVING total_bookings > 0
        ORDER BY total_revenue DESC
        LIMIT 50
    """)
    top_tradespeople = cursor.fetchall()

    # ---- Subscription breakdown ----
    cursor.execute("""
        SELECT sp.plan_name,
               sp.monthly_price,
               COUNT(s.subscription_id) AS active_subscribers,
               COALESCE(SUM(s.price_at_purchase), 0) AS mrr_contribution
        FROM subscription_plans sp
        LEFT JOIN subscriptions s
               ON s.plan_id = sp.plan_id AND s.status='active'
        GROUP BY sp.plan_id, sp.plan_name, sp.monthly_price
        ORDER BY sp.monthly_price ASC
    """)
    sub_breakdown = cursor.fetchall()

    cursor.close(); db.close()

    # ============================================================
    # Build workbook
    # ============================================================
    wb = Workbook()

    # ---- Sheet 1: Summary ----
    ws = wb.active
    ws.title = "Summary"
    ws.append(['TradeConnect Revenue Report'])
    ws.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    ws.append([])
    ws.append(['Metric', 'Value'])
    ws.append(['Total Payment Volume', total_payment_volume])
    ws.append(['Total Platform Fees', total_platform_fees])
    ws.append(['Active Subscriptions', active_subscriptions])
    ws.append(['Monthly Subscription Revenue', monthly_subscription_revenue])
    ws.append(['Total Estimated Revenue', total_estimated_revenue])
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 18

    # ---- Sheet 2: Monthly Revenue ----
    ws = wb.create_sheet("Monthly Revenue")
    ws.append([
        'Month', 'Number of Payments', 'Total Revenue', 'Platform Fees',
        'Avg Payment Size', 'Unique Employers', 'Unique Tradespeople'
    ])
    for row in monthly_data:
        num_payments = int(row['num_payments'])
        total_revenue = float(row['total_revenue'])
        avg_payment = total_revenue / num_payments if num_payments else 0
        ws.append([
            str(row['month']),
            num_payments,
            total_revenue,
            float(row['platform_fees']),
            round(avg_payment, 2),
            int(row['unique_employers']),
            int(row['unique_tradespeople']),
        ])
    for col, width in zip('ABCDEFG', [12, 22, 16, 16, 18, 18, 22]):
        ws.column_dimensions[col].width = width

    # ---- Sheet 3: Revenue by City ----
    ws = wb.create_sheet("Revenue by City")
    ws.append([
        'City', 'Bookings', 'Revenue', 'Platform Fees',
        'Avg Booking Value', 'Active Tradespeople'
    ])
    for row in city_data:
        bookings = int(row['bookings'])
        revenue = float(row['revenue'])
        avg_booking = revenue / bookings if bookings else 0
        ws.append([
            str(row['city'] or ''),
            bookings,
            revenue,
            float(row['platform_fees']),
            round(avg_booking, 2),
            int(row['active_tradespeople']),
        ])
    for col, width in zip('ABCDEF', [22, 12, 14, 16, 20, 22]):
        ws.column_dimensions[col].width = width

    # ---- Sheet 4: Top Tradespeople ----
    ws = wb.create_sheet("Top Tradespeople")
    ws.append([
        'Name', 'Trade Category', 'City', 'Total Bookings',
        'Total Revenue', 'Average Rating'
    ])
    for row in top_tradespeople:
        ws.append([
            str(row['name']),
            str(row['trade_category'] or ''),
            str(row['city'] or ''),
            int(row['total_bookings']),
            float(row['total_revenue']),
            round(float(row['avg_rating']), 2),
        ])
    for col, width in zip('ABCDEF', [26, 18, 18, 16, 16, 16]):
        ws.column_dimensions[col].width = width

    # ---- Sheet 5: Subscription Breakdown ----
    ws = wb.create_sheet("Subscription Breakdown")
    ws.append([
        'Plan Name', 'Monthly Price', 'Active Subscribers',
        'MRR Contribution', '% of Total MRR'
    ])
    total_mrr = sum(float(r['mrr_contribution']) for r in sub_breakdown) or 1
    for row in sub_breakdown:
        mrr = float(row['mrr_contribution'])
        pct = (mrr / total_mrr) * 100 if total_mrr else 0
        ws.append([
            str(row['plan_name']),
            float(row['monthly_price']),
            int(row['active_subscribers']),
            mrr,
            round(pct, 1),
        ])
    for col, width in zip('ABCDE', [16, 16, 20, 18, 16]):
        ws.column_dimensions[col].width = width

    # ============================================================
    # Send file
    # ============================================================
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"tradeconnect_revenue_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
